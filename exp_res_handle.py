import openpyxl
import re

xlsx_path = r'kernelTCP_exp_res.xlsx'

wb = openpyxl.load_workbook(xlsx_path)

# log_path = r'/home/userpath/projects/kernel-tcp/logs/main-2024-03-29-16:21:44.txt'

# lines = []
# with open(log_path, r'r') as f:
#     for line in f:
#         lines.append(line)
# f.close()


# distance_metrics = [
#     'hanming_distance',
#     'edit_distance',
#     'euclidean_string_distance',
#     'manhattan_string_distance',
#     'normalized_compression_distance'
# ]

# for idx in range(len(distance_metrics)):
#     distance_metric = distance_metrics[idx]
#     sheet.cell(1, idx + 2, distance_metric)
#     history_infos = []
#     for line_idx in range(len(lines)):
#         cur_line = lines[line_idx]
#         if distance_metric in cur_line and r'commit' in cur_line:
#             commit_match = re.search(r'commit: (.*?),', cur_line)
#             apfd_match = re.search(r'apfd: (.*?)\n', cur_line)
#             git_sha = commit_match.group(1)
#             apfd = apfd_match.group(1)
#             history_infos.append((git_sha, apfd))
    
#     for version_idx in range(len(history_infos)):
#         hi = history_infos[version_idx]
#         sheet.cell(version_idx + 2, idx + 2, hi[1])
#         sheet.cell(version_idx + 2, 1, hi[0])

# log_path = r'/home/userpath/projects/kernel-tcp/logs/main-2024-04-03-23:42:26.txt'

# lines = []
# with open(log_path, r'r') as f:
#     for line in f:
#         lines.append(line)
# f.close()
# for line_idx in range(len(lines)):
#     sheet.cell(1, 7, 'FAST')
#     history_infos = []
#     for line_idx in range(len(lines)):
#         cur_line = lines[line_idx]
#         if r'commit' in cur_line:
#             commit_match = re.search(r'commit: (.*?),', cur_line)
#             apfd_match = re.search(r'apfd: (.*?)\n', cur_line)
#             git_sha = commit_match.group(1)
#             apfd = apfd_match.group(1)
#             history_infos.append((git_sha, apfd))
    
#     for version_idx in range(len(history_infos)):
#         hi = history_infos[version_idx]
#         sheet.cell(version_idx + 2, 7, hi[1])

# log_path = r'/home/userpath/projects/kernel-tcp/logs/main-2024-03-28-12:40:04.txt'

# lines = []
# with open(log_path, r'r') as f:
#     for line in f:
#         lines.append(line)
# f.close()

# for line_idx in range(len(lines)):
#     sheet.cell(1, 8, 'TOPIC MODEL')
#     history_infos = []
#     for line_idx in range(len(lines)):
#         cur_line = lines[line_idx]
#         if r'topic_main.py : INFO  topic_model distance_metric: jensen_shannon_distance' in cur_line:
#             # commit_match = re.search(r'commit: (.*?),', cur_line)
#             apfd_match = re.search(r'apfd (.*?): (.*?)\n', cur_line)
#             git_sha = apfd_match.group(1)
#             apfd = apfd_match.group(2)
#             history_infos.append((git_sha, apfd))
    
#     for version_idx in range(len(history_infos)):
#         hi = history_infos[version_idx]
#         sheet.cell(version_idx + 2, 8, hi[1])

# log_path = r'/home/userpath/projects/kernel-tcp/logs/main-2024-04-08-21:30:11.txt'

# lines = []
# with open(log_path, r'r') as f:
#     for line in f:
#         lines.append(line)
# f.close()
# sheet.cell(1, 3, 'additional')
# history_infos = []
# for line_idx in range(len(lines)):
#     cur_line = lines[line_idx]
#     if r'cg_based.py : DEBUG  stragety: additional, commit:' in cur_line:
#         commit_match = re.search(r'commit: (.*?),', cur_line)
#         apfd_match = re.search(r'apfd: (.*?)\n', cur_line)
#         git_sha = commit_match.group(1)
#         apfd = apfd_match.group(1)
#         history_infos.append((git_sha, apfd))

# for version_idx in range(len(history_infos)):
#     hi = history_infos[version_idx]
#     sheet.cell(version_idx + 2, 1, hi[0])
#     sheet.cell(version_idx + 2, 3, hi[1])

distance_metrics = [
    'hanming_distance',
    'edit_distance',
    'euclidean_string_distance',
    'manhattan_string_distance',
    'normalized_compression_distance'
]

k = 10

candidate_strategies = [
    'min_max'
]

ltp_version = [
    '20230127',
    '20230516',
    '20230929',
    '20240129'
]

log_path = r'/home/userpath/projects/kernel-tcp/logs/main-2024-05-06-22:32:15.txt'

lines = []
with open(log_path, r'r') as f:
    for line in f:
        lines.append(line)
f.close()

sheet = wb[f'CG-flaky']



sheet.cell(1, 2, 'total')
history_infos = []
for line_idx in range(len(lines)):
    cur_line = lines[line_idx]
    if r'cg_based.py : DEBUG  stragety: total, commit: ' in cur_line:
        commit_match = re.search(r'commit: (.*?),', cur_line)
        apfd_match = re.search(r'apfd: (.*?),', cur_line)
        git_sha = commit_match.group(1)
        apfd = apfd_match.group(1)
        history_infos.append((git_sha, apfd))

for version_idx in range(len(history_infos)):
    git_sha = sheet.cell(version_idx + 2, 1).value
    for hi in history_infos:
        if hi[0] != git_sha:
            continue
    # hi = history_infos[version_idx]
        sheet.cell(version_idx + 2, 2, float(hi[1]))
        break

wb.save(xlsx_path)
